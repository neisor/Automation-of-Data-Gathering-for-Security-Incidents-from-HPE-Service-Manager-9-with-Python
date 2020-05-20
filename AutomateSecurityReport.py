#!/usr/bin/env python
#coding: utf-8
#Created by Antonio Raffaele Iannaccone

import pyautogui
import time
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import clipboard
import tkinter
from tkinter.filedialog import askopenfilename
import json

#FailSafe
pyautogui.FAILSAFE = True


#defining global variables for path to the Excel file and images
global filePath
filePath = ''

global image1var
image1var = ''

global image2var
image2var = ''

global image3var
image3var = ''

global image4var
image4var = ''

global image5var
image5var = ''

global imageStatus
imageStatus = ''

#define longer wait variable after searching for the ticket - useful when searching for archived tickets
global alertWindowVariable
alertWindowVariable = 0

try:
    #read and populate the data from JSON config file
    with open('AutomateSecurityReport_config.json', 'r') as f:
        config = json.load(f)

    #Read the keys from JSON and populate the variables of images
    image1var = config['key1']
    image2var = config['key2']
    image3var = config['key3']
    image4var = config['key4']
    image5var = config['key5']
    imageStatus = config['key6']
except:
    print('JSON config file does not exist yet or is empty...')
    pyautogui.alert(title='Information', text='Before using this program, please, choose your reference images.')


#Define choose excel
def chooseExcel():
    global filePath
    filePath = askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")]) # show an "Open" dialog box and return the path to the selected file
    filePath = filePath.replace("/", "\\")
    

#Define main function
def mainFunction():
    global alertWindowVariable

    #Handling saving of the Excel sheet when CTRL+C is pressed
    import signal
    import sys
    def signal_handler(sig, frame):
            print('Saving the excel sheet before exiting the program...')
            wb.save('AutomateSecurityReport_Output.xlsx')
            print('Saved! Exiting...')
            sys.exit(0)
    signal.signal(signal.SIGINT, signal_handler)

    #Handle the waiting time
    try:
        waitTimeVar = int(waitingTimeText.get(1.0, tkinter.END))
    except:
        pyautogui.alert(title='Error', text="Seems like the value of waiting time after searching for a ticket is too small (less than 6 seconds), or, it's empty. Fill it correctly and try again.")
        alertWindowVariable = 1
        return

    if waitTimeVar <=6:
          pyautogui.alert(title='Error', text="Seems like the value of waiting time after searching for a ticket is too small (less than 6 seconds), or, it's empty. Fill it correctly and try again.")
          alertWindowVariable = 1
          return
          

    print('Starting main logic of the function now...')

    #main try of the function
    try:
    
        try:
        #WORKING WITH EXCEL
        #Get the number of rows and columns in the Excel sheet
            print('Path to the Excel spreadsheet is: ' + filePath)
            wb = load_workbook(str(filePath))
            sheet = wb.worksheets[0]
            row_count = sheet.max_row
            column_count = sheet.max_column
        except:
           pyautogui.alert(title='Error', text='Seems like no Excel sheet has been selected. Please, select one...')

        cellNumber = 2
        firstRowIdentificator = 1

        #Get all the values from Excel with Incident IDs
        for cell in sheet['B']:

           if firstRowIdentificator is 1:
              firstRowIdentificator = 2
              continue


           print (cell.value)
           #Identify SMartEase Addon and write into it
           incidentManagement = pyautogui.locateCenterOnScreen(str(image1var),confidence=.70) #SMart Ease Addon icon in SM9
           pyautogui.moveTo(incidentManagement)
           pyautogui.move(70, 0)
           pyautogui.click()
           time.sleep(0.1)
           incidentID = cell.value
           pyautogui.typewrite(incidentID)
           time.sleep(0.2)
           pyautogui.hotkey('enter')
           time.sleep(int(waitingTimeText.get(1.0, tkinter.END)))


           ### NEW SECTION FOR STATUS ###
           statusImageFind = pyautogui.locateCenterOnScreen(str(imageStatus),confidence=.70) #Status section
           pyautogui.moveTo(statusImageFind)
           pyautogui.move(90, 0)
           pyautogui.click()
           time.sleep(0.1)
           pyautogui.hotkey('ctrl', 'a')
           time.sleep(0.2)
           pyautogui.hotkey('ctrl', 'c')
           time.sleep(0.2)
           statusOfIncident = clipboard.paste()
           time.sleep(0.1)
           sheet['J' + str(cellNumber)] = statusOfIncident
           

           #Scroll down to the Updates section in the ticket
           clickForGoingDownOnPage = pyautogui.locateCenterOnScreen(str(image2var),confidence=.70) #Click in the middle of the incident page to allow it to scroll down
           pyautogui.click(clickForGoingDownOnPage)
           pyautogui.scroll(-700)


           #Click on the text in the Updates section and copy it all
           foundSolutionField = 'NoValueYet'
           try:
               updatesFieldInIM = pyautogui.locateCenterOnScreen(str(image3var),confidence=.70) #Updates section
               foundSolutionField = 'No'
           except TypeError:
               solutionFieldInIM = pyautogui.locateCenterOnScreen(str(image4var),confidence=.70)#Solution section
               foundSolutionField = 'Yes'

           if foundSolutionField is 'No':
               pyautogui.moveTo(updatesFieldInIM)
               pyautogui.move(200, 30)
               pyautogui.click()
               time.sleep(0.1)
               pyautogui.hotkey('ctrl', 'a')
               time.sleep(0.2)
               pyautogui.hotkey('ctrl', 'c')
               time.sleep(0.2)
               incidentID = clipboard.paste()
               sheet['I' + str(cellNumber)] = incidentID

           if foundSolutionField is 'Yes':
               sheet['I' + str(cellNumber)] = 'INCIDENT IS RESOLVED'
               incidentID = 'INCIDENT IS RESOLVED'

           cellNumber += 1

           closeTicketButton = pyautogui.locateCenterOnScreen(str(image5var),confidence=.70) #X (close) button in SM9
           pyautogui.click(closeTicketButton)
           time.sleep(1.5)

        print('Done. Saving the excel sheet before exiting the program...')
        wb.save('AutomateSecurityReport_Output.xlsx')
        print('Saved! Exiting...')
        pyautogui.alert(title='Successfully finished', text='The program finished successfully. Please, check the outputted Excel file.')

    except:
        print('Something went wrong! Saving the excel sheet before exiting the program...')
        wb.save('AutomateSecurityReport_Output.xlsx')
        print('Saved! Exiting...')
        pyautogui.alert(title='Error', text='Something went wrong. Probably your SM9 webpage took too long to load, or, you do not even have the SM9 webpage started up in your browser. That happens sometimes. Check the outputted Excel sheet what information are you still missing in there, then, try again.')

        

#Image1 Chooser
def image1():
    global image1var
    global image2var
    global image3var
    global image4var
    global image5var
    global imageStatus
    image1var = askopenfilename()
    global config
    config = {'key1': image1var, 'key2': image2var, 'key3': image3var, 'key4': image4var, 'key5': image5var, 'key6': imageStatus}
    with open('AutomateSecurityReport_config.json', 'w') as f:
        json.dump(config, f)
        
#Image2 Chooser
def image2():
    global image1var
    global image2var
    global image3var
    global image4var
    global image5var
    global imageStatus
    image2var = askopenfilename()
    global config
    config = {'key1': image1var, 'key2': image2var, 'key3': image3var, 'key4': image4var, 'key5': image5var, 'key6': imageStatus}
    with open('AutomateSecurityReport_config.json', 'w') as f:
        json.dump(config, f)
        
#Image3 Chooser
def image3():
    global image1var
    global image2var
    global image3var
    global image4var
    global image5var
    global imageStatus
    image3var = askopenfilename()
    global config
    config = {'key1': image1var, 'key2': image2var, 'key3': image3var, 'key4': image4var, 'key5': image5var, 'key6': imageStatus}
    with open('AutomateSecurityReport_config.json', 'w') as f:
        json.dump(config, f)
        
#Image4 Chooser
def image4():
    global image1var
    global image2var
    global image3var
    global image4var
    global image5var
    global imageStatus
    image4var = askopenfilename()
    global config
    config = {'key1': image1var, 'key2': image2var, 'key3': image3var, 'key4': image4var, 'key5': image5var, 'key6': imageStatus}
    with open('AutomateSecurityReport_config.json', 'w') as f:
        json.dump(config, f)
        
#Image5 Chooser
def image5():
    global image1var
    global image2var
    global image3var
    global image4var
    global image5var
    global imageStatus
    image5var = askopenfilename()
    global config
    config = {'key1': image1var, 'key2': image2var, 'key3': image3var, 'key4': image4var, 'key5': image5var, 'key6': imageStatus}
    with open('AutomateSecurityReport_config.json', 'w') as f:
        json.dump(config, f)

#Status Image Chooser
def statusImageChooser():
    global image1var
    global image2var
    global image3var
    global image4var
    global image5var
    global imageStatus
    imageStatus = askopenfilename()
    global config
    config = {'key1': image1var, 'key2': image2var, 'key3': image3var, 'key4': image4var, 'key5': image5var, 'key6': imageStatus}
    with open('AutomateSecurityReport_config.json', 'w') as f:
        json.dump(config, f)


#GUI        
top = tkinter.Tk()

top.title('Security Report Automation')
top.geometry('640x330')

button = tkinter.Button(text = 'Start!', command = mainFunction)
button.grid(row = 1, column=1, columnspan=6)

buttonChoose = tkinter.Button(text = 'Choose Excel', command = chooseExcel)
buttonChoose.grid(row = 2, column=1, columnspan=6)

label = tkinter.Label(text = 'Type the amount of time (in seconds) that you want this program to wait\nafter searching for an incident (standard is 9 seconds):')
label.grid(row = 3, column=1, columnspan=6)

waitingTimeText = tkinter.Text(height = 1, width = 5)
waitingTimeText.grid(row = 4, column=1, columnspan=6)

label = tkinter.Label(text = '\nCreated by')
label.grid(row = 5, column=1, columnspan=6)

label2 = tkinter.Label(text = 'Antonio Raffaele Iannaccone\nantonio-raffaele.iannaccone@t-systems.com', borderwidth=2, relief="ridge")
label2.grid(row = 6, column=1, columnspan=6)

label3 = tkinter.Label(text = '')
label3.grid(row = 7, column=1, columnspan=6)

label4 = tkinter.Label(text = 'Use the following buttons to set up reference images:')
label4.grid(row = 8, column=1, columnspan=6)

button1 = tkinter.Button(text = 'AddOn Image', command = image1)
button1.grid(row = 9, column=1, columnspan=1)

button2 = tkinter.Button(text = 'Prioritization image', command = image2)
button2.grid(row = 9, column=2, columnspan=1)

button3 = tkinter.Button(text = 'Updates image', command = image3)
button3.grid(row = 9, column=3, columnspan=1)

button4 = tkinter.Button(text = 'Solution image', command = image4)
button4.grid(row = 9, column=4, columnspan=1)

button5 = tkinter.Button(text = 'X image', command = image5)
button5.grid(row = 9, column=5, columnspan=1)

button6 = tkinter.Button(text = 'Status image', command = statusImageChooser)
button6.grid(row = 9, column=6, columnspan=1)

top.mainloop()

